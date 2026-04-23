import subprocess
import sys

from openpyxl import load_workbook

def run_solution_excel(file_name: str):
    exlce_file = load_workbook(
        filename=file_name,
        read_only=True
    )
    for num_test, row in enumerate(
        iterable=exlce_file.active.iter_rows(min_row=2, values_only=True), 
        start=1
    ):
        input, output = row
        if not input or not output:
            continue
        
        try:
            result = subprocess.run(
                args=[sys.executable, 'solution.py'],
                input=input,
                text=True,
                capture_output=True,
                check=True
            ).stdout.strip()

            if result != output:
                return f'Ошибка в тесте #{num_test}\nВвод:\n{input}\nОжидаемый вывод:\n{output}\nПолученный вывод:\n{result}'
            
        except subprocess.CalledProcessError as e:
            only_error= e.stderr[e.stderr.find('line'):]
            return f'При запуске программы возникла ошибка: {only_error}'

    return 'Все тесты пройдены!'

print(run_solution_excel(file_name='spo_python_sem1_task1.xlsx'))