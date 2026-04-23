from openpyxl import load_workbook

file = load_workbook(
    filename='spo_python_sem1_task1.xlsx',
    read_only=True
)

for row in file.active.iter_rows(min_row=2, values_only=True):
    input, output = row
    if not input or not output:
        continue
    print(input, output)